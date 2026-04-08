import argparse
from pathlib import Path
import sys

import openpyxl


def normalize(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    return value


def to_sql_value(value):
    if value is None:
        return "NULL"
    if isinstance(value, (int, float)):
        return str(value)
    escaped = str(value).replace("'", "''")
    return f"'{escaped}'"


def main():
    parser = argparse.ArgumentParser(description="Generate SQL inserts from the Products Import sheet.")
    parser.add_argument("xlsx", type=Path, help="Path to the XLSX file (Products Import sheet).")
    parser.add_argument("--sheet", default="Products Import", help="Sheet name to read.")
    parser.add_argument("--out", type=Path, help="Output SQL file path (defaults to stdout).")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[args.sheet]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        raise SystemExit("Sheet doesn't look like a products import template.")

    header_row = rows[3]
    columns = [normalize(c) for c in header_row]

    expected = [
        "sku", "product_name", "category", "sub_category", "brand",
        "unit", "reorder_level", "description", "hsn_code", "notes",
    ]
    normalized = [str(c).split()[0].lower() if c else "" for c in columns]
    if normalized[:10] != expected:
        print("Warning: Header columns do not match expected template.", file=sys.stderr)

    inserts = []
    for row in rows[4:]:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        data = {
            "sku": normalize(row[0]),
            "product_name": normalize(row[1]),
            "category": normalize(row[2]),
            "sub_category": normalize(row[3]),
            "brand": normalize(row[4]),
            "unit": normalize(row[5]) or "pcs",
            "reorder_level": normalize(row[6]) or 0,
            "description": normalize(row[7]),
            "hsn_code": normalize(row[8]),
            "notes": normalize(row[9]),
        }
        if not data["sku"] or not data["product_name"] or not data["category"]:
            continue
        columns_sql = ", ".join(data.keys())
        values_sql = ", ".join(to_sql_value(v) for v in data.values())
        inserts.append(
            "insert into products (id, {cols}, price, stock_qty, status, created_at, updated_at) "
            "values (gen_random_uuid()::text, {values}, 0, 0, 'OUT_OF_STOCK', now(), now());"
            .format(cols=columns_sql, values=values_sql)
        )

    output = "\n".join(inserts)
    if args.out:
        args.out.write_text(output, encoding="utf-8")
    else:
        print(output)


if __name__ == "__main__":
    main()
